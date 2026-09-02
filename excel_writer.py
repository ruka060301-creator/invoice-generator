"""
請求書自動作成アプリ - Excel書き込みモジュール

Excelテンプレートを開き、貼り付けシートにデータを書き込み、
請求書シートの数式とヘッダー情報を更新して保存する。
"""

import calendar
import datetime
import os

import openpyxl

import config


class ExcelWriteError(Exception):
    """Excelの書き込みに失敗したときに投げる例外。"""


def create_invoice(template_path, df, year, month, output_dir):
    """テンプレートに予約データを流し込み、請求書ファイルを作成する。

    Args:
        template_path: Excelテンプレートのパス
        df: csv_loader が返した DataFrame
        year:  対象年(例: 2026)
        month: 対象月(例: 8)
        output_dir: 出力先フォルダ

    Returns:
        str: 作成したファイルのパス

    Raises:
        ExcelWriteError: 書き込みに失敗した場合
    """
    record_count = len(df)
    if record_count > config.MAX_RECORDS:
        raise ExcelWriteError(
            f"予約が {record_count} 件あり、テンプレートの上限 "
            f"{config.MAX_RECORDS} 件を超えています。\n"
            "テンプレートの集計範囲を広げる必要があります。"
        )

    workbook = _load_template(template_path)
    paste_sheet = _find_paste_sheet(workbook)
    hotel_name = _extract_hotel_name(paste_sheet.title)

    _clear_paste_sheet(paste_sheet)
    _write_paste_sheet(paste_sheet, df)

    invoice_sheet = _get_invoice_sheet(workbook)
    _rebuild_detail_formulas(invoice_sheet, paste_sheet.title, record_count)
    _update_header(invoice_sheet, year, month, hotel_name)

    return _save(workbook, output_dir, year, month, hotel_name)


# ---------------------------------------------------------------------------
# テンプレートを開く / シートを探す
# ---------------------------------------------------------------------------

def _load_template(template_path):
    """テンプレートを開く。

    data_only を指定しない(=デフォルトの False)ことで、既存の数式が
    文字列として保持される。data_only=True で開いて保存すると、
    すべての数式が計算結果の数値に置き換わって壊れてしまう。
    """
    if not os.path.exists(template_path):
        raise ExcelWriteError(f"テンプレートが見つかりません:\n{template_path}")
    try:
        return openpyxl.load_workbook(template_path)
    except Exception as e:
        raise ExcelWriteError(f"テンプレートを開けませんでした。\n詳細: {e}")


def _find_paste_sheet(workbook):
    """「〜貼り付けシート」という名前のシートを探す。

    シート名に施設名が含まれる(例:「ランズコンドホテル貼り付けシート」)ため、
    完全一致ではなく末尾一致で探すことで、どの施設のテンプレートでも動く。
    """
    for sheet in workbook.worksheets:
        if sheet.title.endswith(config.PASTE_SHEET_SUFFIX):
            return sheet
    raise ExcelWriteError(
        f"「{config.PASTE_SHEET_SUFFIX}」で終わる名前のシートが見つかりません。"
    )


def _get_invoice_sheet(workbook):
    """請求書シートを取得する。"""
    if config.INVOICE_SHEET_NAME not in workbook.sheetnames:
        raise ExcelWriteError(
            f"「{config.INVOICE_SHEET_NAME}」シートが見つかりません。"
        )
    return workbook[config.INVOICE_SHEET_NAME]


def _extract_hotel_name(sheet_title):
    """シート名から施設名を取り出す。

    「ランズコンドホテル貼り付けシート」→「ランズコンドホテル」
    """
    return sheet_title[: -len(config.PASTE_SHEET_SUFFIX)]


# ---------------------------------------------------------------------------
# 貼り付けシートへの書き込み
# ---------------------------------------------------------------------------

def _clear_paste_sheet(sheet):
    """貼り付けシートの既存データを消す(1行目のヘッダーは残す)。

    先に消しておかないと、前月のほうが件数が多かった場合に古い行が残り、
    今月の請求額に前月分が混ざってしまう。
    """
    if sheet.max_row < config.PASTE_START_ROW:
        return
    for row in sheet.iter_rows(
        min_row=config.PASTE_START_ROW,
        max_row=sheet.max_row,
    ):
        for cell in row:
            cell.value = None


def _write_paste_sheet(sheet, df):
    """DataFrame の内容を貼り付けシートに書き込む。"""
    for offset, (_, record) in enumerate(df.iterrows()):
        row = config.PASTE_START_ROW + offset
        for column_name, column_index in config.PASTE_COLUMN_MAP.items():
            sheet.cell(row=row, column=column_index, value=record[column_name])


# ---------------------------------------------------------------------------
# 請求書シートの更新
# ---------------------------------------------------------------------------

def _rebuild_detail_formulas(sheet, paste_sheet_title, record_count):
    """明細行の数式を件数分だけ作り直す。

    テンプレートには前回作成時の件数分の数式しか入っていないため、
    毎回すべて作り直す。余った行はクリアして、前回の数式が残らないようにする。
    """
    quoted = _quote_sheet_name(paste_sheet_title)

    for index in range(config.MAX_RECORDS):
        row = config.DETAIL_START_ROW + index
        source_row = config.PASTE_START_ROW + index  # 参照先の貼り付けシートの行

        if index < record_count:
            site_cell = f"{_column_letter(config.DETAIL_COL_SITE)}{row}"
            amount_ref = f"{quoted}!Z{source_row}"

            sheet.cell(row=row, column=config.DETAIL_COL_CHECKIN,
                       value=f"={quoted}!E{source_row}")
            sheet.cell(row=row, column=config.DETAIL_COL_CHECKOUT,
                       value=f"={quoted}!F{source_row}")
            sheet.cell(row=row, column=config.DETAIL_COL_NIGHTS,
                       value=f"={quoted}!H{source_row}")
            sheet.cell(row=row, column=config.DETAIL_COL_SITE,
                       value=f"={quoted}!I{source_row}")
            sheet.cell(row=row, column=config.DETAIL_COL_ROOM_TYPE,
                       value=f"={quoted}!J{source_row}")
            # Agoda は手数料が引かれた金額で入ってくるため割り戻す
            sheet.cell(
                row=row,
                column=config.DETAIL_COL_AMOUNT,
                value=(
                    f'=IF({site_cell}="{config.ADJUST_SITE_NAME}",'
                    f"{amount_ref}/{config.ADJUST_DIVISOR},{amount_ref})"
                ),
            )
        else:
            # 件数を超えた行は空にする(前回の数式を残さない)
            for column in range(config.DETAIL_COL_CHECKIN,
                                config.DETAIL_COL_AMOUNT + 1):
                sheet.cell(row=row, column=column, value=None)


def _update_header(sheet, year, month, hotel_name):
    """請求書の年月に関する項目を更新する。"""
    # 「8月送客手数料　ランズコンドホテル」(区切りは全角スペース)
    sheet[config.CELL_COMMISSION_LABEL] = f"{month}月送客手数料\u3000{hotel_name}"

    # 請求書番号は「2026-08」形式
    sheet[config.CELL_INVOICE_NUMBER] = f"{year}-{month:02d}"

    # 発行日は翌月1日、支払期限は翌月末日
    issue_date = _next_month_first_day(year, month)
    sheet[config.CELL_ISSUE_DATE] = issue_date
    sheet[config.CELL_DUE_DATE] = _last_day_of_month(
        issue_date.year, issue_date.month
    )


# ---------------------------------------------------------------------------
# 保存
# ---------------------------------------------------------------------------

def _save(workbook, output_dir, year, month, hotel_name):
    """出力先フォルダに保存し、そのパスを返す。"""
    filename = config.OUTPUT_FILENAME_PATTERN.format(
        year=year, month=month, hotel=hotel_name
    )
    output_path = os.path.join(output_dir, filename)

    try:
        os.makedirs(output_dir, exist_ok=True)
        workbook.save(output_path)
    except PermissionError:
        raise ExcelWriteError(
            f"ファイルを保存できませんでした。\n{filename} を Excel で開いていませんか?"
        )
    except Exception as e:
        raise ExcelWriteError(f"保存に失敗しました。\n詳細: {e}")

    return output_path


# ---------------------------------------------------------------------------
# 小さな補助関数
# ---------------------------------------------------------------------------

def _quote_sheet_name(name):
    """数式内で使うシート名。スペースなどを含む場合に備えて必ず引用符で囲む。"""
    escaped = name.replace("'", "''")
    return f"'{escaped}'"


def _column_letter(column_index):
    """列番号を列記号に変換する(3 → 'C')。"""
    return openpyxl.utils.get_column_letter(column_index)


def _next_month_first_day(year, month):
    """翌月1日を返す。"""
    if month == 12:
        return datetime.date(year + 1, 1, 1)
    return datetime.date(year, month + 1, 1)


def _last_day_of_month(year, month):
    """その月の末日を返す。"""
    last_day = calendar.monthrange(year, month)[1]
    return datetime.date(year, month, last_day)